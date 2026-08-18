import sqlite3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io
import base64
import sys
import os

def export_equipments():
    db_path = "c:/Users/TOPIC/Desktop/ftth/ftth.db"
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # 1. Fetch all connections
    cursor.execute("SELECT * FROM connections")
    connections = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    # 2. Replicate the grouping logic from React
    olts = {}
    fdts = {}
    bpis = {}
    
    for c in connections:
        # OLT
        if c.get("port_olt"):
            olt_name = f"OLT Port {c['port_olt']}"
            if olt_name not in olts:
                olts[olt_name] = {
                    "type": "OLT",
                    "name": olt_name,
                    "parent": "Central Office Soukra",
                    "gps": "—",
                    "subscribers": 0
                }
            olts[olt_name]["subscribers"] += 1
            
        # FDT
        if c.get("fdt"):
            fdt_name = c["fdt"]
            if fdt_name not in fdts:
                fdts[fdt_name] = {
                    "type": "FDT",
                    "name": fdt_name,
                    "parent": c.get("residence") or "SOUKRA",
                    "gps": c.get("gps_fdt") or "36.8643276,10.2167956",
                    "subscribers": 0
                }
            fdts[fdt_name]["subscribers"] += 1
            
        # BPI
        if c.get("pos_bpi"):
            bpi_name = c["pos_bpi"]
            if bpi_name not in bpis:
                bpis[bpi_name] = {
                    "type": "BPI",
                    "name": bpi_name,
                    "parent": c.get("fdt") or "Cabinet",
                    "gps": c.get("gps_bpi") or "36.8671225,10.2253475",
                    "subscribers": 0
                }
            bpis[bpi_name]["subscribers"] += 1
            
    # Combine into a single list
    all_eq = list(olts.values()) + list(fdts.values()) + list(bpis.values())
    
    # 3. Create Excel Workbook using openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equipements FTTH"
    
    # Style definitions
    font_title = Font(name="Segoe UI", size=14, bold=True, color="1e3a8a")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="ffffff")
    font_body = Font(name="Segoe UI", size=10)
    font_bold_body = Font(name="Segoe UI", size=10, bold=True)
    
    fill_header = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    fill_zebra = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='e2e8f0'),
        right=Side(style='thin', color='e2e8f0'),
        top=Side(style='thin', color='e2e8f0'),
        bottom=Side(style='thin', color='e2e8f0')
    )
    
    # Title Row
    ws.merge_cells("A1:E1")
    ws["A1"] = "INVENTAIRE DES EQUIPEMENTS RESEAU FTTH (SOTETEL)"
    ws["A1"].font = font_title
    ws["A1"].alignment = align_left
    ws.row_dimensions[1].height = 30
    
    # Headers
    headers = ["Type d'Equipement", "Nom de l'Equipement", "Parent / Zone", "Coordonnées GPS", "Nombre d'Abonnés"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
    ws.row_dimensions[3].height = 25
    
    # Data Rows
    current_row = 4
    for idx, eq in enumerate(all_eq):
        row_cells = [
            ws.cell(row=current_row, column=1, value=eq["type"]),
            ws.cell(row=current_row, column=2, value=eq["name"]),
            ws.cell(row=current_row, column=3, value=eq["parent"]),
            ws.cell(row=current_row, column=4, value=eq["gps"]),
            ws.cell(row=current_row, column=5, value=eq["subscribers"])
        ]
        
        # Apply body styles
        is_even = (idx % 2 == 0)
        for c_idx, cell in enumerate(row_cells, 1):
            cell.font = font_bold_body if c_idx == 2 else font_body
            cell.border = thin_border
            if is_even:
                cell.fill = fill_zebra
            
            if c_idx in [1, 4, 5]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        
    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue # Skip title row for length check
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    # 4. Save to BytesIO and print base64 string to stdout
    output = io.BytesIO()
    wb.save(output)
    excel_bytes = output.getvalue()
    b64_string = base64.b64encode(excel_bytes).decode("utf-8")
    
    sys.stdout.write(b64_string)

if __name__ == "__main__":
    export_equipments()
